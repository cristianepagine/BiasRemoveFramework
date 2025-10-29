"""
Gerador de Relatórios Excel Formatados
Gera Excel com múltiplas abas, formatação profissional e conditional formatting
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule


class ExcelReportGenerator:
    """
    Gerador de relatórios Excel com formatação profissional.

    Gera Excel com 4 abas:
    1. Resumo Executivo (tabela 3 cenários)
    2. Detecção de Viés (com conditional formatting)
    3. Eficácia da Correção (com % em verde/vermelho)
    4. Mudanças de Posição (com setas ↑↓)
    """

    def __init__(self, output_dir: str = "reports/excel"):
        """
        Inicializa o gerador de Excel.

        Args:
            output_dir: Diretório para salvar os arquivos Excel
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Estilos padrão
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.cell_font = Font(name='Arial', size=10)
        self.cell_alignment = Alignment(horizontal='left', vertical='center')

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _formatar_header(self, ws, row: int = 1):
        """Aplica formatação ao cabeçalho"""
        for cell in ws[row]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _ajustar_largura_colunas(self, ws):
        """Ajusta largura das colunas automaticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def aba_resumo_executivo(
        self,
        wb,
        dados_cenarios: Dict[str, Dict[str, float]]
    ):
        """
        Aba 1: Resumo Executivo

        Args:
            wb: Workbook do openpyxl
            dados_cenarios: Dados dos 3 cenários
                {
                    'Cenário 1 - Sem Correção': {
                        'Média Feminino': 7.2,
                        'Média Masculino': 7.8,
                        'Diferença': 0.6,
                        'P-value': 0.001,
                        'Viés Detectado': 'Sim'
                    },
                    ...
                }
        """
        ws = wb.create_sheet("Resumo Executivo", 0)

        # Título
        ws['A1'] = 'RESUMO EXECUTIVO - ANÁLISE DE VIÉS'
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Linha em branco
        ws.append([])

        # Cabeçalhos
        headers = ['Cenário', 'Média Feminino', 'Média Masculino',
                  'Diferença', 'P-value', 'Viés Detectado']
        ws.append(headers)
        self._formatar_header(ws, row=3)

        # Dados
        for cenario, metricas in dados_cenarios.items():
            row = [
                cenario,
                f"{metricas.get('Média Feminino', 0):.2f}",
                f"{metricas.get('Média Masculino', 0):.2f}",
                f"{metricas.get('Diferença', 0):.3f}",
                f"{metricas.get('P-value', 0):.4f}",
                metricas.get('Viés Detectado', 'N/A')
            ]
            ws.append(row)

        # Formatação das células de dados
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Conditional formatting para P-value
        # Verde se > 0.05 (sem viés), vermelho se < 0.05 (com viés)
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Aplica manualmente baseado no valor
        for row_idx in range(4, ws.max_row + 1):
            p_value_cell = ws.cell(row=row_idx, column=5)
            vies_cell = ws.cell(row=row_idx, column=6)

            try:
                p_val = float(p_value_cell.value)
                if p_val < 0.05:
                    p_value_cell.fill = red_fill
                    vies_cell.fill = red_fill
                    vies_cell.font = Font(name='Arial', size=10, bold=True, color='9C0006')
                else:
                    p_value_cell.fill = green_fill
                    vies_cell.fill = green_fill
                    vies_cell.font = Font(name='Arial', size=10, bold=True, color='006100')
            except:
                pass

        # Adiciona resumo estatístico
        ws.append([])
        ws.append(['ANÁLISE COMPARATIVA'])
        summary_row = ws.max_row
        ws.merge_cells(f'A{summary_row}:F{summary_row}')
        ws[f'A{summary_row}'].font = Font(name='Arial', size=12, bold=True, color='366092')
        ws[f'A{summary_row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        self._ajustar_largura_colunas(ws)

        print("✓ Aba 'Resumo Executivo' criada")

    def aba_deteccao_vies(
        self,
        wb,
        dados_deteccao: pd.DataFrame
    ):
        """
        Aba 2: Detecção de Viés

        Args:
            wb: Workbook do openpyxl
            dados_deteccao: DataFrame com colunas:
                ['Tipo_Avaliacao', 'Genero', 'N_Amostras', 'Media', 'Desvio_Padrao',
                 'Diferenca_Percentual', 'P_value', 'Vies_Detectado']
        """
        ws = wb.create_sheet("Detecção de Viés")

        # Título
        ws['A1'] = 'DETECÇÃO DE VIÉS POR TIPO DE AVALIAÇÃO'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='366092')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])

        # Escreve DataFrame
        for r in dataframe_to_rows(dados_deteccao, index=False, header=True):
            ws.append(r)

        # Formata cabeçalho
        self._formatar_header(ws, row=3)

        # Formata células de dados
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Conditional formatting

        # 1. Diferença Percentual (coluna F)
        # Verde: -5% a +5%, Amarelo: -10% a -5% ou +5% a +10%, Vermelho: < -10% ou > +10%
        for row_idx in range(4, ws.max_row + 1):
            diff_cell = ws.cell(row=row_idx, column=6)
            try:
                # Remove % e converte
                diff_val = float(str(diff_cell.value).replace('%', ''))

                if -5 <= diff_val <= 5:
                    diff_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif -10 <= diff_val < -5 or 5 < diff_val <= 10:
                    diff_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    diff_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            except:
                pass

        # 2. P-value (coluna G)
        # Verde se > 0.05, vermelho se < 0.05
        for row_idx in range(4, ws.max_row + 1):
            p_cell = ws.cell(row=row_idx, column=7)
            try:
                p_val = float(p_cell.value)
                if p_val < 0.05:
                    p_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    p_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            except:
                pass

        # 3. Viés Detectado (coluna H)
        for row_idx in range(4, ws.max_row + 1):
            vies_cell = ws.cell(row=row_idx, column=8)
            if vies_cell.value == 'Sim':
                vies_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                vies_cell.font = Font(name='Arial', size=10, bold=True, color='9C0006')
            elif vies_cell.value == 'Não':
                vies_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                vies_cell.font = Font(name='Arial', size=10, bold=True, color='006100')

        self._ajustar_largura_colunas(ws)

        print("✓ Aba 'Detecção de Viés' criada")

    def aba_eficacia_correcao(
        self,
        wb,
        dados_eficacia: pd.DataFrame
    ):
        """
        Aba 3: Eficácia da Correção

        Args:
            wb: Workbook do openpyxl
            dados_eficacia: DataFrame com colunas:
                ['Tipo_Avaliacao', 'Diferenca_Antes', 'Diferenca_Depois',
                 'Reducao_Absoluta', 'Reducao_Percentual', 'Eficacia']
        """
        ws = wb.create_sheet("Eficácia da Correção")

        # Título
        ws['A1'] = 'EFICÁCIA DA CORREÇÃO DE VIÉS'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='366092')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])

        # Escreve DataFrame
        for r in dataframe_to_rows(dados_eficacia, index=False, header=True):
            ws.append(r)

        # Formata cabeçalho
        self._formatar_header(ws, row=3)

        # Formata células de dados
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Conditional formatting para Redução Percentual (coluna E)
        # Verde: > 50%, Amarelo: 25-50%, Vermelho: < 25%
        for row_idx in range(4, ws.max_row + 1):
            red_cell = ws.cell(row=row_idx, column=5)
            try:
                # Remove % e converte
                red_val = float(str(red_cell.value).replace('%', ''))

                if red_val >= 50:
                    red_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    red_cell.font = Font(name='Arial', size=10, bold=True, color='006100')
                elif 25 <= red_val < 50:
                    red_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    red_cell.font = Font(name='Arial', size=10, bold=True, color='9C5700')
                else:
                    red_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    red_cell.font = Font(name='Arial', size=10, bold=True, color='9C0006')
            except:
                pass

        # Conditional formatting para Eficácia (coluna F)
        for row_idx in range(4, ws.max_row + 1):
            ef_cell = ws.cell(row=row_idx, column=6)
            if ef_cell.value == 'Alta':
                ef_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                ef_cell.font = Font(name='Arial', size=10, bold=True, color='006100')
            elif ef_cell.value == 'Média':
                ef_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                ef_cell.font = Font(name='Arial', size=10, bold=True, color='9C5700')
            elif ef_cell.value == 'Baixa':
                ef_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                ef_cell.font = Font(name='Arial', size=10, bold=True, color='9C0006')

        self._ajustar_largura_colunas(ws)

        print("✓ Aba 'Eficácia da Correção' criada")

    def aba_mudancas_posicao(
        self,
        wb,
        dados_mudancas: pd.DataFrame
    ):
        """
        Aba 4: Mudanças de Posição

        Args:
            wb: Workbook do openpyxl
            dados_mudancas: DataFrame com colunas:
                ['Pessoa_ID', 'Nome', 'Genero', 'Posicao_Antes', 'Posicao_Depois',
                 'Mudanca', 'Direcao']
        """
        ws = wb.create_sheet("Mudanças de Posição")

        # Título
        ws['A1'] = 'MUDANÇAS DE POSIÇÃO NO RANKING'
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color='366092')
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append([])

        # Escreve DataFrame
        for r in dataframe_to_rows(dados_mudancas, index=False, header=True):
            ws.append(r)

        # Formata cabeçalho
        self._formatar_header(ws, row=3)

        # Formata células de dados
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adiciona setas na coluna Mudanca (coluna F)
        for row_idx in range(4, ws.max_row + 1):
            mudanca_cell = ws.cell(row=row_idx, column=6)
            direcao_cell = ws.cell(row=row_idx, column=7)

            try:
                mudanca = int(mudanca_cell.value)
                if mudanca > 0:
                    mudanca_cell.value = f"↑ {mudanca}"
                    mudanca_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    mudanca_cell.font = Font(name='Arial', size=11, bold=True, color='006100')
                    direcao_cell.value = "Subiu"
                    direcao_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif mudanca < 0:
                    mudanca_cell.value = f"↓ {abs(mudanca)}"
                    mudanca_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    mudanca_cell.font = Font(name='Arial', size=11, bold=True, color='9C0006')
                    direcao_cell.value = "Desceu"
                    direcao_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                else:
                    mudanca_cell.value = "→ 0"
                    mudanca_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    direcao_cell.value = "Manteve"
                    direcao_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            except:
                pass

        self._ajustar_largura_colunas(ws)

        print("✓ Aba 'Mudanças de Posição' criada")

    def gerar_relatorio_completo(
        self,
        dados_completos: Dict,
        nome_arquivo: Optional[str] = None
    ) -> Path:
        """
        Gera relatório Excel completo com todas as abas.

        Args:
            dados_completos: Dicionário com todos os dados:
                {
                    'resumo_executivo': Dict com dados dos cenários,
                    'deteccao_vies': DataFrame,
                    'eficacia_correcao': DataFrame,
                    'mudancas_posicao': DataFrame
                }
            nome_arquivo: Nome do arquivo (opcional)

        Returns:
            Path do arquivo gerado
        """
        if nome_arquivo is None:
            nome_arquivo = f"relatorio_vies_{self.timestamp}.xlsx"

        caminho = self.output_dir / nome_arquivo

        print("\n=== Gerando Relatório Excel ===\n")

        # Cria workbook
        wb = openpyxl.Workbook()

        # Remove a planilha padrão
        wb.remove(wb.active)

        # Cria abas
        try:
            if 'resumo_executivo' in dados_completos:
                self.aba_resumo_executivo(wb, dados_completos['resumo_executivo'])
        except Exception as e:
            print(f"⚠ Erro ao criar aba Resumo Executivo: {e}")

        try:
            if 'deteccao_vies' in dados_completos:
                self.aba_deteccao_vies(wb, dados_completos['deteccao_vies'])
        except Exception as e:
            print(f"⚠ Erro ao criar aba Detecção de Viés: {e}")

        try:
            if 'eficacia_correcao' in dados_completos:
                self.aba_eficacia_correcao(wb, dados_completos['eficacia_correcao'])
        except Exception as e:
            print(f"⚠ Erro ao criar aba Eficácia da Correção: {e}")

        try:
            if 'mudancas_posicao' in dados_completos:
                self.aba_mudancas_posicao(wb, dados_completos['mudancas_posicao'])
        except Exception as e:
            print(f"⚠ Erro ao criar aba Mudanças de Posição: {e}")

        # Salva arquivo
        wb.save(caminho)

        print(f"\n✓ Relatório Excel salvo: {caminho}")

        return caminho
